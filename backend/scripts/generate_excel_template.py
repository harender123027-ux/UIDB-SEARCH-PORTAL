import re
import sqlite3

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation


def sanitize_name(name):
    """Excel named ranges cannot have spaces or special characters."""
    clean = re.sub(r'[^a-zA-Z0-9]', '_', name)
    return f"DIST_{clean}"

def generate_excel_template(db_path, output_path):
    # 1. Fetch data from DB
    conn = sqlite3.connect(db_path)
    districts = pd.read_sql_query("SELECT DISTINCT name FROM districts WHERE is_active = 1 ORDER BY name ASC", conn)['name'].tolist()
    stations_df = pd.read_sql_query("""
        SELECT d.name as district, ps.name as station
        FROM police_stations ps
        JOIN districts d ON ps.district_id = d.id
        WHERE ps.is_active = 1
        ORDER BY d.name, ps.name ASC
    """, conn)
    conn.close()

    # 2. Setup Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UI Body Template"

    # Exact Attributes from App (17 data + 10 image slots = 27); matches bulk_import_ui_bodies.py
    headers = [
        "dd_no", "found_date", "found_district", "ps_name", "found_loc",
        "gender", "age_min", "age_max", "height_cm", "build",
        "skin_tone", "hair_color", "beard", "visible_marks",
        "clothing_description", "notes", "additional_details",
        "image_face_frontal_path", "image_face_left_path",
        "image_face_right_path", "image_full_body_path",
        "image_tattoo_1_path", "image_tattoo_2_path", "image_tattoo_3_path",
        "image_tattoo_4_path", "image_tattoo_5_path", "image_tattoo_6_path",
        "image_tattoo_7_path", "image_tattoo_8_path", "image_tattoo_9_path",
        "image_tattoo_10_path",
        "image_clothing_path", "image_belonging_path",
    ]
    ws.append(headers)

    # 3. Reference Data Sheet
    ref_ws = wb.create_sheet("ReferenceData")

    # Static Lists from bis-pwa.jsx
    gender_list = ["Male", "Female", "Unknown"]
    build_list = ["Slim", "Medium", "Heavy", "Unknown"]
    skin_list = ["Fair", "Medium", "Dark", "Unknown"]
    hair_list = ["Black", "Grey", "Brown", "White", "Unknown"]
    beard_list = ["Yes", "No", "N/A"]

    for col, lst in enumerate([gender_list, build_list, skin_list, hair_list, beard_list], start=1):
        for i, v in enumerate(lst):
            ref_ws.cell(row=i + 1, column=col, value=v)

    # Write Districts (Col F)
    for i, d in enumerate(districts):
        ref_ws.cell(row=i+1, column=6, value=d)

    wb.create_named_range("DistrictList", ref_ws, f"$F$1:$F${len(districts)}")

    # Write Stations per District (Starting Col G)
    col_idx = 7
    for district in districts:
        dist_stations = stations_df[stations_df['district'] == district]['station'].tolist()
        if not dist_stations:
            continue

        ref_ws.cell(row=1, column=col_idx, value=district)
        for row_idx, station in enumerate(dist_stations):
            ref_ws.cell(row=row_idx+2, column=col_idx, value=station)

        sanitized = sanitize_name(district)
        col_letter = get_column_letter(col_idx)
        wb.create_named_range(sanitized, ref_ws, f"${col_letter}$2:${col_letter}${len(dist_stations)+1}")
        col_idx += 1

    # 4. Add Data Validations (Matches App order)

    # District (Col C)
    dv_district = DataValidation(type="list", formula1="DistrictList", allow_blank=True)
    ws.add_data_validation(dv_district)
    dv_district.add("C2:C2000")

    # PS Name (Col D) - Cascading
    formula_indirect = 'INDIRECT("DIST_"&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(C2," ","_"),"-","_"),".","_"),"(","_"),")","_"))'
    dv_ps = DataValidation(type="list", formula1=formula_indirect, allow_blank=True)
    ws.add_data_validation(dv_ps)
    dv_ps.add("D2:D2000")

    # Gender (Col F)
    dv_gender = DataValidation(type="list", formula1=f"{quote_sheetname('ReferenceData')}!$A$1:$A$3", allow_blank=True)
    ws.add_data_validation(dv_gender)
    dv_gender.add("F2:F2000")

    # Build (Col J)
    dv_build = DataValidation(type="list", formula1=f"{quote_sheetname('ReferenceData')}!$B$1:$B$4", allow_blank=True)
    ws.add_data_validation(dv_build)
    dv_build.add("J2:J2000")

    # Skin Tone (Col K)
    dv_skin = DataValidation(type="list", formula1=f"{quote_sheetname('ReferenceData')}!$C$1:$C$4", allow_blank=True)
    ws.add_data_validation(dv_skin)
    dv_skin.add("K2:K2000")

    # Hair Color (Col L)
    dv_hair = DataValidation(type="list", formula1=f"{quote_sheetname('ReferenceData')}!$D$1:$D$5", allow_blank=True)
    ws.add_data_validation(dv_hair)
    dv_hair.add("L2:L2000")

    # Beard (Col M)
    dv_beard = DataValidation(type="list", formula1=f"{quote_sheetname('ReferenceData')}!$E$1:$E$3", allow_blank=True)
    ws.add_data_validation(dv_beard)
    dv_beard.add("M2:M2000")

    # 5. Styling
    from openpyxl.styles import Alignment, Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    ws.freeze_panes = "A2"
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    # Instructions Sheet
    inst_ws = wb.create_sheet("Instructions")
    inst_ws.append(["App Attribute Alignment Guide"])
    inst_ws.append(["This template matches the 'New Case' form in the UBIS app exactly."])
    inst_ws.append(["Columns R-AA: Use relative file paths for images (e.g., images/case1/face.jpg)"])
    inst_ws.append(["Tattoo columns (U-Z, AA): Support up to 10 tattoo/mark/person items; fill in as many as available"])

    ref_ws.sheet_state = 'hidden'
    wb.save(output_path)
    print(f"App-aligned Excel template created at: {output_path}")

if __name__ == "__main__":
    db = "/Users/anmoldureha/Code/FaceRecognitionSystem/backend/ubis.db"
    output = "/Users/anmoldureha/Code/FaceRecognitionSystem/ui_body_template.xlsx"
    generate_excel_template(db, output)
